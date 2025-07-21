from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import openpyxl as excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
import time
import os



bills_robot = webdriver.Chrome()
bills_robot.maximize_window()
bills_robot.get(" Web App Link ")
hold = WebDriverWait( bills_robot ,20)

gear_box = ActionChains(bills_robot)

account_passed = ""

def scrolling():
    """Control Scroll for all"""
    bills_robot.execute_script("window.scrollTo(0,400);")


def spinner() :
    WebDriverWait(bills_robot, 20).until(EC.presence_of_element_located((By.CLASS_NAME, "CLASS_NAME")))


def login_page(username,password) :
    """Login FX"""
    # Check Point To Move.
    hold.until(EC.visibility_of_element_located((By.XPATH,"//div[@id='name']//p")))

    # Enter UserName
    hold.until(EC.visibility_of_element_located((By.ID,"username"))).send_keys(username)
    # Enter Password
    hold.until(EC.visibility_of_element_located((By.ID, "password"))).send_keys(password)

    # Step Tab + Space
    gear_box.send_keys(Keys.TAB).send_keys(Keys.SPACE).perform()

    # Press Customers Button
    hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Customers"))).click()
    time.sleep(0.25)


def amount_cycle(acc):
    global account_passed
    invoices_list = [f"Account {str(acc)}" , "" , "" , "" , "" , "" , "" , "", 0]

    try :
        def page_source(account):
            """Enter account number to pass source page"""

            # Press Search to view source fields
            hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search"))).click()

            # select from Drop Menu
            drop_menu = hold.until(EC.element_to_be_clickable((By.CLASS_NAME, "CLASS_NAME")))
            choice = Select(drop_menu)
            choice.select_by_index(0)

            # Enter The Account Number
            hold.until(EC.visibility_of_element_located((By.ID, "ID"))).send_keys(account)

            # Press Search.
            hold.until(EC.element_to_be_clickable((By.ID, "ID"))).click()

            # Pres to enter the account
            hold.until(EC.element_to_be_clickable((By.CLASS_NAME, "CLASS_NAME"))).click()


        page_source(acc)
        spinner()

        time.sleep(0.30)
        hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Billing and rating"))).click()
        spinner()

        time.sleep(0.20)
        hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Financial overview"))).click()
        spinner()

        open_items_radio = hold.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='XPATH'][1]")))
        if not open_items_radio.is_selected() :
            open_items_radio.click()
            spinner()

        time.sleep(0.35)
        hold.until(EC.element_to_be_clickable((By.ID, "ID name"))).click()
        spinner()

        scrolling()

        time.sleep(0.35)
        invoices_table = hold.until(EC.visibility_of_all_elements_located((By.XPATH, "//table[@id='Name']/tbody[2]/tr")))
        # What if not found ?


        total_bills = 0


        for i in  invoices_table :
            row_type = i.find_element(By.XPATH, "//table[@id='name']/tbody[2]/tr[1]/td[6]").text
            if row_type != "Invoice" :
                continue

            else:
                bill_month = i.find_element(By.XPATH, ".//td[5]").text

                bl_amount_str = i.find_element(By.XPATH, ".//td[8]").text

                bl_amount_int = ""

                for ch in bl_amount_str :
                    if ch != "E" :
                        bl_amount_int += ch

                    elif ch == "E" :
                        break


                cleaned = bl_amount_int.strip().replace(",","")
                bl_amount_float = float(cleaned)
                total_bills += bl_amount_float



                if "Jan" in bill_month :
                    invoices_list[1] = bl_amount_float

                if "Feb" in bill_month :
                    invoices_list[2] = bl_amount_float

                if "Mar" in bill_month :
                    invoices_list[3] = bl_amount_float

                if "Apr" in bill_month :
                    invoices_list[4] = bl_amount_float

                if "May" in bill_month :
                    invoices_list[5] = bl_amount_float

                if "Jun" in bill_month :
                    invoices_list[6] = bl_amount_float



        invoices_list[7] = len(invoices_table)
        invoices_list[8] = total_bills


        hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search"))).click()
        time.sleep(0.3)




    except Exception as er :
        account_passed = False
        print(er)
        time.sleep(0.3)
        bills_robot.save_screenshot(f"Bills Error_{acc}.png")
        time.sleep(0.3)
        bills_robot.refresh()

        try:
            spinner()
        except :
            hold.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='footer']//p")))
        finally:
            time.sleep(0.5)

        hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search"))).click()


    else:
        account_passed = True


    return invoices_list







# Start Engine Sequence
sequence = 0
# File Handling || Source Phase--1
source_file_path = r" Path .xlsx"
source_wb = excel.load_workbook(source_file_path)
source_sheet = source_wb["accounts list"]


# File Handling || Destination Phase--2
destination_file_path = r" Path .xlsx"
destination_wb = excel.load_workbook(destination_file_path)
destination_sheet = destination_wb["Sheet1"]



# Login FX
login_page("User Name "," Password ")


for profile in range( 2 , source_sheet.max_row+1 ) :

    account_number = str(source_sheet.cell(profile, 1).value)

    sequence += 1
    print(f"Handling Account : {account_number} || Sequence : {sequence}")

    # List Value
    result_full_account_data = amount_cycle(account_number)


    account_status = source_sheet.cell(profile, 2)
    if account_passed == True:
        account_status.value = "Done"
    else:
        account_status.value = "Not Done"

    # Save in source sheet.
    source_wb.save(source_file_path)


    # Append the List to Excel file
    destination_sheet.append(result_full_account_data)
    destination_wb.save(destination_file_path)




print("Saving Data....")
time.sleep(2)
bills_robot.quit()
time.sleep(2)
print("Task has been performed.")